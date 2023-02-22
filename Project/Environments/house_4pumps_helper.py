import pandas as pd
import numpy as np
import pytz
import pickle
import os
from datetime import date, datetime, timedelta, tzinfo, time
import openpyxl


class HPutils:
    def __init__(self, SensiboDevices: dict, dataPath: str) -> None:
        self.SensiboDevices = SensiboDevices

    def CreatePumpState(self, TimeSchedule: datetime, sensibo_data) -> dict:
        State = {}

        for HP in self.SensiboDevices:
            df_sensibo_pump = pd.json_normalize(sensibo_data[HP], sep='_')
            df_sensibo_pump = df_sensibo_pump.explode(list(df_sensibo_pump.columns), ignore_index=True)
            dateMask = (df_sensibo_pump['time'] <= TimeSchedule)

            CurrentState = {
                'targetTemperature': df_sensibo_pump[dateMask].iloc[-1]['states_targetTemperature'],
                'fanLevel': df_sensibo_pump[dateMask].iloc[-1]['states_fanLevel'],
                'on': df_sensibo_pump[dateMask].iloc[-1]['states_on'],
                'mode': df_sensibo_pump[dateMask].iloc[-1]['states_mode']
            }
            State[HP] = CurrentState

        return State

    def CreatePumpDataStructures(self) -> dict:
        HPstates = ['targetTemperature', 'fanLevel', 'on', 'mode']
        HPmeas = ['temperature', 'humidity']

        # Creates structures with initial lists
        Data = {}
        for key in self.SensiboDevices.keys():
            Data[key] = {'flattime': [],
                         'Time': [],
                         'measurements': {},
                         'states': {}}

            for keymeas in HPmeas:
                Data[key]['measurements'][keymeas] = []
            for keystate in HPstates:
                Data[key]['states'][keystate] = []

        return Data

    def GetPumpState(self, TimeSchedule: datetime, sensibo_data) -> dict:
        State = {}

        for HP in self.SensiboDevices:
            df_sensibo_pump = pd.json_normalize(sensibo_data[HP], sep='_')
            df_sensibo_pump = df_sensibo_pump.explode(list(df_sensibo_pump.columns), ignore_index=True)
            dateMask = (df_sensibo_pump['time'] <= TimeSchedule)

            CurrentState = {
                # 'targetTemperature': df_sensibo_pump[dateMask].iloc[-1]['states_targetTemperature'],
                'Fan': df_sensibo_pump[dateMask].iloc[-1]['states_fanLevel'],
                'On': df_sensibo_pump[dateMask].iloc[-1]['states_on'],
                'mode': df_sensibo_pump[dateMask].iloc[-1]['states_mode'],
                'Room': df_sensibo_pump[dateMask].iloc[-1]['measurements_temperature']
            }

            Fandict = {'quiet': 1, 'low': 2, 'medium': 3, 'medium_high': 4, 'high': 5}

            CurrentState['Fan'] = Fandict[CurrentState['Fan']]
            State[HP] = CurrentState

        return State

    def EstimateWallTemperature(self, RoomTemp: float, OutTemp: float, rho_in: float, rho_out: float):
        '''
        Wall Temperature is estimated as steady state 1D heat conduction through Wall
        Inputs:
        RoomTemp: Room Temperature
        OutTemp: Outside Temperature
        rho_in: heat conduction parameter for inside wall
        rho_out: heat conduction parameter for outside wall

        Output
        WallTemp: Wall Temperature
        '''
        WallTemp = (rho_in * RoomTemp + rho_out * OutTemp) / (rho_in + rho_out)

        return WallTemp

    def EstimateHPPower(self, GroupPower, pumps):
        '''
        Placeholder function, could be replaced by either MHE or a steady state power equations
        '''
        n_pumps = len(pumps)
        power = GroupPower / n_pumps

        return power


class DataReader:
    def __init__(self, dataPath, HPutils, config) -> None:

        local_timezone: tzinfo = pytz.timezone('Europe/Oslo')
        Tax = 0.25
        self.HP = HPutils
        self.local_timezone = local_timezone
        self.Tax = Tax
        self.dataPath = dataPath

        ### Load large data files
        self.spot_data = pickle.load(open(os.path.abspath(os.path.join(self.dataPath, 'SpotData2021.pkl')), "rb"))
        self.MET_data = pickle.load(open(os.path.abspath(os.path.join(self.dataPath, 'Data_MET.pkl')), "rb"))
        self.sensibo_data = pickle.load(open(os.path.abspath(os.path.join(self.dataPath, 'Data_Sensibo.pkl')), "rb"))
        # self.tibberPump_data = pickle.load(open(os.path.join(self.dataPath, 'Data_TibberPump.pkl'), "rb"))
        # self.df_tibberPump = pd.DataFrame.from_dict(self.tibberPump_data)
        # self.df_tibberPump_resampled = self.resample_tibber_data(self.df_tibberPump, 5)

        ### Load SYSID data
        self.SYSID_data = pickle.load(open(os.path.join(self.dataPath, 'HouseSYSID_HP.pkl'), "rb"))

        self.PowHistogram = pickle.load(open(os.path.join(self.dataPath, 'PowerHistogramRelative.pkl'), 'rb'))
        self.TempHistogram = pickle.load(open(os.path.join(self.dataPath, 'TempHistogram.pkl'), 'rb'))
        # self.PowHistogram = None
        # self.TempHistogram = None

        self.timesteps = config['timesteps']
        self.timesteps_N = config['timesteps+N']

        self.dt = config['dt']
        self.dt_N = config['dt+N']

    def GetSpotData(self, start, stop):
        data = pd.DataFrame.from_dict(self.spot_data)

        data["Time_start"] = data["Time_start"].dt.tz_convert(self.local_timezone)  # Local time zone
        mask = (data["Time_start"] >= start) & (data["Time_start"] <= stop)

        data = data.loc[mask]
        data = data.drop(columns='Time_end')

        data = data.rename(columns={"Time_start": "Time", "Price": "Prices"})

        prices = data.to_dict(orient='list')

        return prices

    def resample_tibber_data(self, dfTibber: pd.DataFrame, newSamplingTime: int) -> pd.DataFrame:

        newSamplingTime = newSamplingTime * 60  # Convert to seconds

        df = dfTibber.copy()

        df['cum_sampling_time'] = df['sampling_time'].cumsum()

        s = df.groupby(
            (df['cum_sampling_time'] % newSamplingTime == 0).shift(fill_value=0).cumsum()
        )['power'].transform('mean')
        df['avg_power'] = np.where(df['cum_sampling_time'] % newSamplingTime == 0, s, 0)

        df['power'] = df['avg_power']
        df['sampling_time'] = newSamplingTime

        df = df[df['cum_sampling_time'] % newSamplingTime == 0].reset_index(drop=True)

        return df  # df[df.columns[:-2]]

    def get_target_temperature(self, config) -> dict:
        TargetTemperature = {}

        for HP in config['SensiboDevices']:
            TargetTemperature[HP] = []
            df_sensibo_pump = pd.json_normalize(self.sensibo_data[HP], sep='_')
            df_sensibo_pump = df_sensibo_pump.explode(list(df_sensibo_pump.columns), ignore_index=True)

            for TimeSchedule in self.timesteps:
                dateMask = (df_sensibo_pump['time'] <= TimeSchedule)
                TargetTemperature[HP].append(df_sensibo_pump[dateMask].iloc[-1]['states_targetTemperature'])
                # only append Temp, without time
        return TargetTemperature

    # def get_tibber_flattime(self, start: datetime):
    #
    #     df_tibber = self.df_tibberPump.copy()
    #
    #     # Force correct data type and timezone
    #     df_tibber['time'] = pd.to_datetime(df_tibber['time'], errors='coerce', utc=1).dt.tz_convert(
    #         self.local_timezone)
    #
    #     tibberFlattime = []
    #     dt = self.timesteps[0]
    #     dateMask = (df_tibber['time'] == dt)
    #     for _, row in df_tibber[dateMask].iterrows():
    #         tibberFlattime.append((row['time'] - start).total_seconds() / 60.)
    #
    #     return tibberFlattime

    def get_initial_state(self, timesteps, pumps: list, start: datetime, outTemp: float, param: dict,
                          TargetTemp: dict) -> dict:

        state0 = self.HP.GetPumpState(start, self.sensibo_data)  # get "Fan, On, Room, Mode"
        # dateMask = (self.df_tibberPump_resampled['time'] <= start)
        # totalPower = self.df_tibberPump_resampled[dateMask].iloc[-1]['avg_power']

        ### match data unit [W] to simulation unit [kW]
        # totalPower = totalPower*0.001

        for pump in pumps:
            state0[pump]['Room'] = 18
            state0[pump]['Wall'] = self.HP.EstimateWallTemperature(
                state0[pump]['Room'], outTemp, param['rho_in'][pump], param['rho_out'][pump])
            state0[pump]['Power'] = self.HP.EstimateHPPower(1, pumps)

            state0[pump]['TargetTemp'] = TargetTemp[pump][timesteps.index(start)]
            state0[pump]['Slack'] = 0
            state0[pump]['SlackMinTemp'] = 0

        return state0

    # def get_target_temperature(self, SensiboDevices:list, timesteps: list) -> dict:
    #     TargetTemp = {}
    #     for ts in timesteps:
    #         TargetTemp[ts] = {}
    #         for HP in SensiboDevices:
    #             df_sensibo_pump = pd.json_normalize(self.sensibo_data[HP], sep='_')
    #             df_sensibo_pump = df_sensibo_pump.explode(list(df_sensibo_pump.columns), ignore_index=True)
    #             dateMask = (df_sensibo_pump['time'] <= ts)
    #
    #             TargetTemp[ts][HP] = df_sensibo_pump[dateMask].iloc[-1]['states_targetTemperature']
    #     return TargetTemp

    def get_met_data(self, start: datetime, stop: datetime):

        met = pd.DataFrame.from_dict(self.MET_data)

        ### convert timezones
        met['time'] = met['time'].dt.tz_convert(self.local_timezone)

        ### rename and drop columns
        met = met.drop(columns=['sampling_time', 'time_abs', 'time_start'])
        met = met.rename(columns={'time': 'Time', 'temperature': 'Temperature'})

        mask = (met["Time"] >= start) & (met["Time"] <= stop)
        data = met.loc[mask]
        met_data = data.to_dict(orient='list')

        return met_data

    def read_settings(self):
        ### Read desired temperatures
        settings = openpyxl.load_workbook(os.path.join(self.dataPath, 'settings/TempSettings.xlsx'))
        settings = settings.active

        TempSettings = {}
        for column in settings.iter_cols(1, settings.max_column):
            TempSettings[column[0].value] = []

        for i, row in enumerate(settings.iter_rows(values_only=True)):
            if i > 0:
                for colnum, colname in enumerate(TempSettings.keys()):
                    TempSettings[colname].append(row[colnum])

        ### Read minimum temperatures
        settings = openpyxl.load_workbook(os.path.join(self.dataPath, 'settings/MinTempSettings.xlsx'))
        settings = settings.active

        MinTempSettings = {}
        for column in settings.iter_cols(1, settings.max_column):
            MinTempSettings[column[0].value] = []

        for i, row in enumerate(settings.iter_rows(values_only=True)):
            if i > 0:
                for colnum, colname in enumerate(MinTempSettings.keys()):
                    MinTempSettings[colname].append(row[colnum])

        ### Read max fan
        settings = openpyxl.load_workbook(os.path.join(self.dataPath, 'settings/FanMax.xlsx'))
        settings = settings.active

        FanMax = {}
        for column in settings.iter_cols(1, settings.max_column):
            FanMax[column[0].value] = []

        for i, row in enumerate(settings.iter_rows(values_only=True)):
            if i > 0:
                for colnum, colname in enumerate(FanMax.keys()):
                    if not (row[colnum] is None):
                        FanMax[colname].append(row[colnum])

        return TempSettings, MinTempSettings, FanMax

    def get_input_dictionary(self, config) -> dict:
        InputDict = {}
        InputDict['pumps'] = config['SensiboDevices']
        InputDict['COP'] = (4, 0.1, 10)
        InputDict['Timesteps'] = self.timesteps

        ### get house params and scale
        InputDict['params'] = self.SYSID_data['ParamScaled']
        InputDict['params'].pop('PowFreeze')  # drop PowFreeze from dictionary, since it is unused
        InputDict['scale'] = self.SYSID_data['Scale']

        InputDict['SpotPrice'] = self.GetSpotData(config['timesteps+N'][0], config['timesteps+N'][-1])
        InputDict['MET'] = self.get_met_data(config['timesteps+N'][0], config['timesteps+N'][-1])
        InputDict['TargetTemp'] = self.get_target_temperature(config)
        InputDict['state0'] = self.get_initial_state(config['timesteps'], config['SensiboDevices'], config['Start'],
                                                     InputDict['MET']['Temperature'][0], InputDict['params'],
                                                     InputDict['TargetTemp'])
        InputDict['PowerHistogram'] = self.PowHistogram
        InputDict['TempHistogram'] = self.TempHistogram

        TempSettings, MinTempSettings, FanMax = self.read_settings()
        ### Interpolate outside temperature for simulation
        MET_time = []
        for ts in InputDict['MET']['Time']:
            delta = ts - InputDict['MET']['Time'][0]
            MET_time.append(delta.total_seconds() / 60)  # [0, 60, 120...]
            # use [0,5,10..] interpolate
        InputDict['OutTemp'] = list(np.interp(config['dt+N'], MET_time, InputDict['MET']['Temperature']))

        ### Interpolate spot prices for simulation
        Spot_time = []
        for ts in InputDict['SpotPrice']['Time']:
            delta = ts - InputDict['SpotPrice']['Time'][0]
            Spot_time.append(delta.total_seconds() / 60)  # [0, 60, 120...]
            # use [0,5,10..] interpolate
        InputDict['Spot'] = list(np.interp(config['dt+N'], Spot_time, InputDict['SpotPrice']['Prices']))

        ### Interpolate DesiredTemp, MinTem, and MaxFan for simulation
        Setting_time = []
        for ts in config['timesteps+N']:
            Setting_time.append(int(ts.hour + ts.minute / 60))
        InputDict['DesiredTemp'] = {}
        InputDict['MinTemp'] = {}
        InputDict['MaxFan'] = {}
        for pump in InputDict['pumps']:
            InputDict['DesiredTemp'][pump] = list(np.interp(Setting_time, np.array(TempSettings['Times']),
                                                            np.array(TempSettings[pump])))
            InputDict['MinTemp'][pump] = list(np.interp(Setting_time, np.array(MinTempSettings['Times']),
                                                        np.array(MinTempSettings[pump])))
            InputDict['MaxFan'][pump] = list(np.interp(Setting_time, np.array(FanMax['Times']),
                                                       np.array(FanMax[pump])))

        return InputDict


